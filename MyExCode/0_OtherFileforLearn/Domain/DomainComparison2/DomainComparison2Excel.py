import openpyxl
import sqlite3
import os


def Createtable():
    conn = sqlite3.connect("domainName.db")
    sql = '''Create table Domain(DomainName text,NextDate text,Price text,Date text,Sheet text)'''
    conn.execute(sql)
    conn.close()


def AddRecord(SourceFile, dbFileName):

    for sheets in SourceFile.sheetnames:
        # 讀取每個資料表
        sheet = SourceFile[sheets]
        # 紀錄至資料庫
        for row in range(2, sheet.max_row + 1):
            conn = sqlite3.connect(dbFileName)
            if sheet['A' + str(row)].value is not None:
                DomainName = sheet['A' + str(row)].value
                NextDate = sheet['B' + str(row)].value
                Price = sheet['C' + str(row)].value
                Date = sheet['D' + str(row)].value
            DBvalues = (DomainName, NextDate, Price, Date, sheets)
            sql = 'insert into Domain values(?,?,?,?,?)'
            conn.execute(sql, DBvalues)
            conn.commit()
        conn.close()


def SearchRecord(DomainComparisonFile_Obj, dbFileName):
    domainlist = []
    for domain in DomainComparisonFile_Obj:
        try:
            conn2 = sqlite3.connect(dbFileName)
            sql2 = 'SELECT * from Domain where DomainName = \"' + domain + '\"'
            results = conn2.execute(sql2)
            domainlist.append(results.fetchall()[0])
        except Exception:
            print("%s 域名出現錯誤" % domain)
    return domainlist


# 將目前工作目錄設定為指令搞所在目錄
os.chdir(os.path.dirname(os.path.abspath(__file__)))
dbFileName = 'domainName.db'
SourceFile = openpyxl.load_workbook('Source.xlsx')  # 開啟excel檔案
DomainComparisonFile = 'A.txt'
# 第一次執行 建立資料庫、各式文檔
if os.path.exists(dbFileName) is False:
    Createtable()
    AddRecord(SourceFile, dbFileName)
    newfile = open('A.txt', 'w')
    newfile.close()
    print('再啟動一次程式')
else:
    with open(DomainComparisonFile, encoding='utf-8-sig') as file_Obj:
        DomainComparisonFile_Obj = file_Obj.read().splitlines()
    results = SearchRecord(DomainComparisonFile_Obj, dbFileName)  # 列印出結果
    # 輸出結果到文字檔
    with open('Results.txt', 'a', encoding='utf-8') as fileObj:
        for result in results:
            fileObj.write("%s\t%s\t%s\t%s\t分頁名稱：\t%s\n" % (
                result[0], result[1], result[2], result[3], result[4]))

while True:
    flag = input("輸入E關閉程式:")
    if flag.upper() == 'E':
        break
